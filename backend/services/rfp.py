"""
RFP Question Bank generator.

Curates a prioritized set of PBM RFP questions based on contract analysis,
then exports as an Excel (.xlsx) file using openpyxl.
"""

import io
import json
import os
import re
from datetime import date
from typing import Any

import anthropic
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import PBMAnalysisReport


# ── Master Question Bank (122 questions) ──────────────────────────────────────

QUESTION_BANK: list[dict[str, Any]] = [
    # Pricing – Brand Drugs
    {"id": 1,  "category": "Pricing – Brand Drugs", "question": "Provide guaranteed AWP discount percentages for brand drugs dispensed at retail (30-day supply).", "evaluation": "Market benchmark: 15–22% off AWP. Higher discounts reduce plan costs.", "weight": 10},
    {"id": 2,  "category": "Pricing – Brand Drugs", "question": "Provide guaranteed AWP discount percentages for brand drugs dispensed through mail order (90-day supply).", "evaluation": "Market benchmark: 20–28% off AWP for mail. Mail should be meaningfully better than retail.", "weight": 9},
    {"id": 3,  "category": "Pricing – Brand Drugs", "question": "Are brand AWP discounts guaranteed at point-of-sale, or are they post-claims reconciliation?", "evaluation": "Point-of-sale guarantees are preferred. Reconciliation creates cash-flow risk.", "weight": 8},
    {"id": 4,  "category": "Pricing – Brand Drugs", "question": "Describe how you define AWP and which pricing database you reference (e.g., Medi-Span, Red Book).", "evaluation": "Named, auditable pricing database required. Proprietary databases are a red flag.", "weight": 7},
    {"id": 5,  "category": "Pricing – Brand Drugs", "question": "Are DAW (Dispense As Written) brand claims subject to the same AWP discount as non-DAW brand claims?", "evaluation": "DAW claims should receive the same discount. Separate, lesser discounts for DAW claims are unfavorable.", "weight": 7},
    {"id": 6,  "category": "Pricing – Brand Drugs", "question": "Provide the minimum guaranteed brand AWP discount for each plan year over the contract term.", "evaluation": "Multi-year guarantees should show year-over-year improvement. Flat rates are below market.", "weight": 8},
    {"id": 7,  "category": "Pricing – Brand Drugs", "question": "Are brand drugs in limited distribution networks excluded from the AWP discount guarantee?", "evaluation": "Limited distribution drug carve-outs can expose significant specialty spend. Seek inclusion or narrow exclusions.", "weight": 7},
    {"id": 8,  "category": "Pricing – Brand Drugs", "question": "How are brand drugs defined? Describe any carve-outs or exceptions to the brand pricing guarantee.", "evaluation": "Fewer carve-outs is better. Any exceptions should be narrowly defined.", "weight": 6},
    {"id": 9,  "category": "Pricing – Brand Drugs", "question": "What AWP discount applies to brand drugs dispensed at retail under 90-day-at-retail programs?", "evaluation": "90-day retail should receive mail-comparable pricing. Standard retail rates for 90-day supply is below market.", "weight": 7},

    # Pricing – Generic Drugs
    {"id": 10, "category": "Pricing – Generic Drugs", "question": "Provide guaranteed AWP discount percentages for generic drugs dispensed at retail (30-day supply).", "evaluation": "Market benchmark: 78–88% off AWP for retail generics. Under 74% is significantly below market.", "weight": 10},
    {"id": 11, "category": "Pricing – Generic Drugs", "question": "Provide guaranteed AWP discount percentages for generic drugs dispensed through mail order (90-day supply).", "evaluation": "Market benchmark: 80–90% off AWP for mail generics.", "weight": 9},
    {"id": 12, "category": "Pricing – Generic Drugs", "question": "Do you use a Maximum Allowable Cost (MAC) list for generic pricing? If so, describe the MAC methodology.", "evaluation": "MAC lists should be transparent, regularly updated, and subject to appeal rights.", "weight": 9},
    {"id": 13, "category": "Pricing – Generic Drugs", "question": "Will you provide a copy of your MAC list and commit to advance notice before MAC price changes?", "evaluation": "Best practice: full MAC list disclosure + 30-day advance notice of changes.", "weight": 8},
    {"id": 14, "category": "Pricing – Generic Drugs", "question": "Describe the appeals process for MAC pricing decisions. What is the typical resolution time?", "evaluation": "Best practice: formal appeal process, resolution within 5 business days, retroactive adjustment if upheld.", "weight": 8},
    {"id": 15, "category": "Pricing – Generic Drugs", "question": "Are generic AWP discounts guaranteed at point-of-sale, or subject to post-claims reconciliation?", "evaluation": "Point-of-sale guarantees are preferred. Reconciliation creates administrative burden.", "weight": 7},
    {"id": 16, "category": "Pricing – Generic Drugs", "question": "What is the effective generic dispensing rate (GDR) for a plan of our size and therapeutic mix?", "evaluation": "Higher GDR means more savings. Ask for modeling based on actual claim history.", "weight": 7},
    {"id": 17, "category": "Pricing – Generic Drugs", "question": "Are brand drugs that become generic ('first-fill generics') subject to the generic AWP discount guarantee?", "evaluation": "First-fill generics should receive generic-tier pricing. Brand pricing on these is unfavorable.", "weight": 6},
    {"id": 18, "category": "Pricing – Generic Drugs", "question": "How are authorized generic drugs priced — at brand or generic AWP discounts?", "evaluation": "Authorized generics should receive generic-tier pricing. Brand pricing on authorized generics is a spread pricing tactic.", "weight": 8},

    # Pricing – Specialty Drugs
    {"id": 19, "category": "Pricing – Specialty Drugs", "question": "Provide guaranteed AWP discount percentages for specialty drugs dispensed through your specialty pharmacy.", "evaluation": "Market range: 10–20% off AWP for specialty, though this varies significantly by drug class.", "weight": 10},
    {"id": 20, "category": "Pricing – Specialty Drugs", "question": "Are specialty drugs required to be dispensed exclusively through your specialty pharmacy? Describe the network exclusivity policy.", "evaluation": "Exclusive specialty networks can limit member access. Prefer open or limited-exclusive networks with continuity exceptions.", "weight": 9},
    {"id": 21, "category": "Pricing – Specialty Drugs", "question": "Provide specialty drug rebate pass-through guarantees for the top 20 specialty drugs by spend.", "evaluation": "Specialty rebates are significant (often $500–$2,000 PMPY for high-cost plans). Full pass-through is best practice.", "weight": 10},
    {"id": 22, "category": "Pricing – Specialty Drugs", "question": "How do you handle limited distribution drugs (LDDs) that are only available through manufacturer-designated pharmacies?", "evaluation": "LDD carve-outs from specialty discount guarantees should be disclosed and narrowly defined.", "weight": 8},
    {"id": 23, "category": "Pricing – Specialty Drugs", "question": "Describe your biosimilar substitution policies and any savings guarantees when switching from reference biologics.", "evaluation": "Biosimilar programs should include guaranteed savings share. Look for automatic substitution provisions.", "weight": 8},
    {"id": 24, "category": "Pricing – Specialty Drugs", "question": "How do you handle GLP-1 and weight loss medication costs? Are these carve-outs to specialty pricing guarantees?", "evaluation": "GLP-1 drugs are the fastest-growing cost driver. Ensure they are included in discount and rebate guarantees.", "weight": 9},
    {"id": 25, "category": "Pricing – Specialty Drugs", "question": "What clinical management programs do you have for specialty drugs (e.g., PA, step therapy, quantity limits)?", "evaluation": "Robust clinical management can reduce specialty trend. Ask for trend data by therapy class.", "weight": 7},
    {"id": 26, "category": "Pricing – Specialty Drugs", "question": "Provide specialty trend data for plans of comparable size over the past 3 years.", "evaluation": "Compare to industry trend (typically 15–25% annually). Outliers indicate weak clinical management.", "weight": 7},

    # Rebates & Pass-Through
    {"id": 27, "category": "Rebates & Pass-Through", "question": "Provide guaranteed minimum rebate amounts (PMPY) for brand and specialty drugs, by plan year.", "evaluation": "Market range for mid-size groups: $150–$400 PMPY for brand. Guarantees should increase year-over-year.", "weight": 10},
    {"id": 28, "category": "Rebates & Pass-Through", "question": "Is your rebate model pass-through or spread pricing? If spread, disclose the spread retained.", "evaluation": "Pass-through is best practice. Spread pricing means the PBM retains a portion of rebates — quantify the impact.", "weight": 10},
    {"id": 29, "category": "Rebates & Pass-Through", "question": "What percentage of manufacturer rebates are passed through to the plan sponsor?", "evaluation": "Best practice: 100% pass-through. Any retained percentage should be offset by lower admin fees.", "weight": 10},
    {"id": 30, "category": "Rebates & Pass-Through", "question": "Are performance guarantees tied to rebate amounts? Describe any conditions that could reduce rebate payments.", "evaluation": "Rebate guarantees should be unconditional or tied to achievable, measurable thresholds.", "weight": 8},
    {"id": 31, "category": "Rebates & Pass-Through", "question": "What is the rebate payment timeline? How frequently are rebates paid and what is the settlement lag?", "evaluation": "Best practice: quarterly payments with 90-day or less settlement lag. Annual-only is below market.", "weight": 7},
    {"id": 32, "category": "Rebates & Pass-Through", "question": "Provide an itemized breakdown of how rebates are calculated (manufacturer vs. PBM-retained vs. plan-received).", "evaluation": "Full transparency is best practice. Lack of itemization suggests spread pricing.", "weight": 8},
    {"id": 33, "category": "Rebates & Pass-Through", "question": "Are specialty drug rebates included in the rebate guarantee or managed separately?", "evaluation": "Specialty rebates should be included in the guarantee. Separate management often means lower pass-through.", "weight": 8},
    {"id": 34, "category": "Rebates & Pass-Through", "question": "How are inflation rebates (CPI adjustments from manufacturers) handled and passed through to the plan?", "evaluation": "CPI adjustments from manufacturers should be fully passed through to the plan sponsor.", "weight": 7},
    {"id": 35, "category": "Rebates & Pass-Through", "question": "Are administrative fees offset against rebates or charged separately? Provide examples for our group size.", "evaluation": "Bundled fee structures can obscure true net costs. Require separate disclosure of fees and rebates.", "weight": 8},

    # Network & Pharmacy Access
    {"id": 36, "category": "Network & Pharmacy Access", "question": "Provide the total number of network pharmacies nationally, by state, and in our geographic area.", "evaluation": "Adequate access: at least 90% of plan members within 5 miles of a network pharmacy.", "weight": 9},
    {"id": 37, "category": "Network & Pharmacy Access", "question": "What are the dispensing fees paid to network pharmacies? Are they disclosed to the plan sponsor?", "evaluation": "Full fee disclosure is best practice. Higher pharmacy fees reduce plan savings.", "weight": 7},
    {"id": 38, "category": "Network & Pharmacy Access", "question": "Describe your preferred pharmacy network, including eligibility criteria and any differential pricing.", "evaluation": "Preferred networks can drive savings but should not sacrifice member access or choice.", "weight": 8},
    {"id": 39, "category": "Network & Pharmacy Access", "question": "What is your policy on out-of-network pharmacy claims? Are emergency provisions included?", "evaluation": "Seek emergency out-of-network access with standard cost-sharing. Strict network limits are member-unfriendly.", "weight": 6},
    {"id": 40, "category": "Network & Pharmacy Access", "question": "Describe your mail order pharmacy capabilities, average turnaround time, and member satisfaction score.", "evaluation": "Best practice: 3–5 day turnaround, NCQA accreditation, 90%+ member satisfaction.", "weight": 7},
    {"id": 41, "category": "Network & Pharmacy Access", "question": "How do you handle 90-day retail vs. mail order pricing and utilization management?", "evaluation": "Plans should capture mail-order savings. Ensure 90-day retail pricing is not comparable to mail to preserve the mail incentive.", "weight": 7},
    {"id": 42, "category": "Network & Pharmacy Access", "question": "What is your policy on specialty pharmacy network exclusivity and member access exceptions?", "evaluation": "Seek exceptions for continuity of care, rare disease, and member preference with cost-parity requirements.", "weight": 8},

    # Clinical Programs
    {"id": 43, "category": "Clinical Programs", "question": "Describe your prior authorization (PA) program — which drugs require PA and what are the medical criteria?", "evaluation": "PA criteria should be evidence-based and consistent with major clinical guidelines.", "weight": 8},
    {"id": 44, "category": "Clinical Programs", "question": "Describe your step therapy programs and the therapeutic categories covered.", "evaluation": "Step therapy is an important cost management tool. Ensure adequate appeal rights and emergency overrides.", "weight": 7},
    {"id": 45, "category": "Clinical Programs", "question": "What medication therapy management (MTM) programs do you offer? Provide outcomes data.", "evaluation": "MTM programs can reduce costs and improve adherence. Ask for ROI data.", "weight": 6},
    {"id": 46, "category": "Clinical Programs", "question": "Describe your formulary design. How many tiers are offered and how is tier placement determined?", "evaluation": "Formulary design impacts member costs and utilization. Ensure formulary is evidence-based.", "weight": 7},
    {"id": 47, "category": "Clinical Programs", "question": "How do you manage opioid utilization and what safety programs are in place?", "evaluation": "Seek programs that comply with CDC guidelines and state PMP (prescription monitoring) integration.", "weight": 7},
    {"id": 48, "category": "Clinical Programs", "question": "Describe your drug adherence programs for maintenance medications (diabetes, hypertension, etc.).", "evaluation": "Better adherence reduces overall healthcare costs. Programs should be evidence-based with measurable outcomes.", "weight": 6},
    {"id": 49, "category": "Clinical Programs", "question": "How do you handle manufacturer copay accumulator programs and their impact on member deductibles?", "evaluation": "Copay accumulator programs (preventing manufacturer coupons from counting toward deductibles) should be disclosed.", "weight": 7},

    # Reporting & Transparency
    {"id": 50, "category": "Reporting & Transparency", "question": "Describe your standard quarterly and annual reporting package. What metrics are included?", "evaluation": "Best practice: trend analysis, GDR, formulary compliance, rebate detail, network performance.", "weight": 8},
    {"id": 51, "category": "Reporting & Transparency", "question": "Will you provide a full claim-level data file in a standard format (e.g., NCPDP)? What is the lag time?", "evaluation": "Full claim-level data access is best practice. Lag should be 30 days or less.", "weight": 9},
    {"id": 52, "category": "Reporting & Transparency", "question": "Do you provide a real-time or near-real-time claims portal for plan administrators?", "evaluation": "Real-time portals enable proactive management. Quarterly-only reporting is below market.", "weight": 7},
    {"id": 53, "category": "Reporting & Transparency", "question": "Will you disclose all revenue sources including manufacturer fees, data sale revenue, and pharmacy margin?", "evaluation": "Full revenue disclosure is best practice and increasingly required by fiduciary standards.", "weight": 9},
    {"id": 54, "category": "Reporting & Transparency", "question": "Describe your spread pricing disclosure practices. Do you use spread pricing in any channel?", "evaluation": "Spread pricing disclosure is legally required in many states. Zero-spread (pass-through) is best practice.", "weight": 9},
    {"id": 55, "category": "Reporting & Transparency", "question": "What is your process for notifying plan sponsors of formulary changes, network changes, or pricing updates?", "evaluation": "Best practice: 90-day advance notice for formulary changes; 30 days for network changes.", "weight": 7},
    {"id": 56, "category": "Reporting & Transparency", "question": "Provide SSAE 18 SOC 1 Type II audit reports from the past 2 years.", "evaluation": "Unqualified opinions are expected. Qualified opinions or refusal to provide are red flags.", "weight": 7},
    {"id": 57, "category": "Reporting & Transparency", "question": "Will you provide a complete fee disclosure itemizing all compensation received in connection with our plan (Section 408(b)(2) style)?", "evaluation": "ERISA-style fee disclosure is best practice and ensures fiduciary compliance. Refusal may create liability.", "weight": 9},

    # Administrative & Financial
    {"id": 58, "category": "Administrative & Financial", "question": "Provide a complete schedule of all administrative fees (PEPM, per-claim, flat monthly, etc.).", "evaluation": "Market benchmark: $0–$3 PEPM for admin. All fees must be disclosed — hidden fees are a red flag.", "weight": 10},
    {"id": 59, "category": "Administrative & Financial", "question": "Are there any implementation fees, termination fees, or conversion fees? If so, provide details.", "evaluation": "Seek waiver of implementation and termination fees for good-faith exits. Conversion penalties > 3 months are unfavorable.", "weight": 8},
    {"id": 60, "category": "Administrative & Financial", "question": "What financial guarantees exist for administrative performance (claims processing accuracy, payment turnaround)?", "evaluation": "Best practice: 99%+ claims processing accuracy, 48-hour turnaround, financial penalties for misses.", "weight": 7},
    {"id": 61, "category": "Administrative & Financial", "question": "Describe your coordination of benefits (COB) and subrogation processes.", "evaluation": "Efficient COB and subrogation recovery directly reduces plan costs. Ask for historical recovery rates.", "weight": 6},
    {"id": 62, "category": "Administrative & Financial", "question": "How do you handle pharmacy audit recoveries? What percentage of recovered amounts are returned to the plan?", "evaluation": "Best practice: 100% of pharmacy audit recoveries returned to the plan. Retained recoveries are a red flag.", "weight": 8},

    # Audit Rights
    {"id": 63, "category": "Audit Rights", "question": "Does the contract grant the plan sponsor the right to conduct independent financial audits of all PBM transactions?", "evaluation": "Unconditional audit rights are best practice and increasingly required. Restrictions are a red flag.", "weight": 10},
    {"id": 64, "category": "Audit Rights", "question": "Describe the scope of audit rights — does it include rebate calculations, pricing guarantees, and spread pricing?", "evaluation": "Audit scope must include rebate calculations, all pricing guarantees, and any retained revenue.", "weight": 10},
    {"id": 65, "category": "Audit Rights", "question": "What is the lookback period for contract audits?", "evaluation": "Best practice: 3–5 year lookback. Less than 2 years is inadequate.", "weight": 8},
    {"id": 66, "category": "Audit Rights", "question": "Will you provide access to source documentation (claim data, manufacturer contracts, etc.) during an audit?", "evaluation": "Source document access is essential for meaningful audits. Refusal is a major red flag.", "weight": 9},
    {"id": 67, "category": "Audit Rights", "question": "If an audit reveals underpayments, describe the remediation process including interest on recovered amounts.", "evaluation": "Best practice: retroactive correction plus interest (prime + 2%) for underpayments.", "weight": 8},
    {"id": 68, "category": "Audit Rights", "question": "What is the process for engaging a third-party auditor? Are there restrictions on auditor selection?", "evaluation": "Plan sponsor should have unrestricted right to engage any qualified independent auditor. PBM-approved auditor lists are a red flag.", "weight": 9},

    # Contract Terms & Legal
    {"id": 69, "category": "Contract Terms & Legal", "question": "Provide the proposed contract term and renewal provisions, including auto-renewal notice periods.", "evaluation": "Prefer 3-year terms with annual performance review. Auto-renewal with 90-day out is standard.", "weight": 8},
    {"id": 70, "category": "Contract Terms & Legal", "question": "What is the termination for convenience notice period? Are there penalties for early termination?", "evaluation": "Best practice: 90-day termination for convenience with no financial penalty. 180+ days is unfavorable.", "weight": 9},
    {"id": 71, "category": "Contract Terms & Legal", "question": "Does the contract include a most-favored nation (MFN) or most-favored customer clause?", "evaluation": "MFN clauses ensure competitive pricing parity. Absence suggests the PBM may offer better terms to comparable clients.", "weight": 8},
    {"id": 72, "category": "Contract Terms & Legal", "question": "Describe indemnification and liability provisions, including caps on liability and exclusions.", "evaluation": "Seek mutual indemnification. PBM liability caps should be adequate to cover potential financial harm.", "weight": 7},
    {"id": 73, "category": "Contract Terms & Legal", "question": "Does the PBM maintain ERISA fiduciary status for any services? Describe the scope.", "evaluation": "Under current regulatory trends, PBMs increasingly face fiduciary obligations. Seek explicit fiduciary acknowledgment.", "weight": 9},
    {"id": 74, "category": "Contract Terms & Legal", "question": "How does the contract address compliance with state PBM reform laws (MAC transparency, spread pricing disclosure, etc.)?", "evaluation": "30+ states have PBM-specific laws. Contract should include a compliance clause covering applicable state laws.", "weight": 8},
    {"id": 75, "category": "Contract Terms & Legal", "question": "Describe dispute resolution procedures — is binding arbitration required or is litigation available?", "evaluation": "Mandatory arbitration clauses can limit plan sponsor rights. Seek the right to litigate material disputes.", "weight": 6},
    {"id": 76, "category": "Contract Terms & Legal", "question": "What are the contract's non-disclosure and confidentiality requirements? Do they limit the plan's ability to share data with consultants?", "evaluation": "Overly broad NDA clauses can prevent plan sponsors from sharing data with advisors — a red flag.", "weight": 7},
    {"id": 77, "category": "Contract Terms & Legal", "question": "Are there exclusivity requirements that prevent the plan from using other pharmacy vendors or direct contracts?", "evaluation": "Exclusivity requirements limit plan flexibility. Seek carve-outs for specialty pharmacy, direct contracts, etc.", "weight": 8},
    {"id": 78, "category": "Contract Terms & Legal", "question": "Describe your data ownership policy — who owns the adjudicated claims data?", "evaluation": "Plan sponsors should own their claims data. PBM data ownership claims are a red flag.", "weight": 9},
    {"id": 79, "category": "Contract Terms & Legal", "question": "Does the contract include a benchmarking clause allowing renegotiation if terms fall below market?", "evaluation": "Benchmarking clauses protect against market degradation over multi-year terms. Absence leaves plan sponsor at risk.", "weight": 7},

    # Performance Guarantees
    {"id": 80, "category": "Performance Guarantees", "question": "Provide a complete schedule of performance guarantees, including at-risk amounts and measurement methodology.", "evaluation": "Best practice: 1–3% of admin fees at risk against clearly defined, measurable metrics.", "weight": 9},
    {"id": 81, "category": "Performance Guarantees", "question": "Are financial performance guarantees backed by amounts held at risk, or are they prospective commitments only?", "evaluation": "Retrospective guarantees with at-risk amounts are more valuable than prospective commitments.", "weight": 8},
    {"id": 82, "category": "Performance Guarantees", "question": "What performance metrics are measured? Include GDR, formulary compliance, member satisfaction.", "evaluation": "At minimum: GDR guarantee, formulary compliance rate, member satisfaction score, mail order penetration.", "weight": 7},
    {"id": 83, "category": "Performance Guarantees", "question": "Is there a minimum guarantee for drug cost trend (medical inflation rate cap)?", "evaluation": "Drug trend guarantees (e.g., trend not to exceed X% per year) provide budget predictability.", "weight": 8},
    {"id": 84, "category": "Performance Guarantees", "question": "If performance guarantees are missed, what is the remedy process and timing for reimbursement?", "evaluation": "Best practice: automatic reimbursement within 60 days of measurement period end.", "weight": 7},

    # Member Experience
    {"id": 85, "category": "Member Experience", "question": "Describe member-facing digital tools, including mobile app, drug cost lookup, and pharmacy finder.", "evaluation": "Modern tools directly impact member satisfaction and proper drug utilization.", "weight": 7},
    {"id": 86, "category": "Member Experience", "question": "What is your member services support model? What are average hold times and hours?", "evaluation": "Best practice: 24/7 support, average hold time under 2 minutes, bilingual support available.", "weight": 7},
    {"id": 87, "category": "Member Experience", "question": "Describe your member appeals and grievance process for coverage denials.", "evaluation": "Seek clear, ACA-compliant appeal processes with defined timelines (72-hour urgent, 30-day standard).", "weight": 7},
    {"id": 88, "category": "Member Experience", "question": "How do you communicate formulary changes to members and provide transition of care support?", "evaluation": "Best practice: 30-day advance notice to affected members with transition supply provisions.", "weight": 6},
    {"id": 89, "category": "Member Experience", "question": "Provide member satisfaction scores (CAHPS or equivalent) for the past 3 years.", "evaluation": "Target: 85%+ member satisfaction. Declining scores may indicate service quality issues.", "weight": 7},
    {"id": 90, "category": "Member Experience", "question": "Describe your specialty pharmacy copay assistance program and manufacturer coupon coordination policies.", "evaluation": "Copay assistance programs reduce member out-of-pocket costs. Copay accumulator programs excluding assistance are member-unfriendly.", "weight": 7},

    # Compliance & Regulatory
    {"id": 91, "category": "Compliance & Regulatory", "question": "Describe your compliance program for federal and state PBM laws, including recent state reforms.", "evaluation": "PBM should have dedicated compliance staff and a documented compliance program.", "weight": 8},
    {"id": 92, "category": "Compliance & Regulatory", "question": "Are you subject to any current or pending regulatory investigations, litigation, or consent decrees?", "evaluation": "Material investigations or litigation should be disclosed and assessed for business risk.", "weight": 8},
    {"id": 93, "category": "Compliance & Regulatory", "question": "How do you address HIPAA and data security requirements? Describe your data breach notification process.", "evaluation": "HIPAA BAA required. Seek Tier 1 data security certification (SOC 2 Type II or equivalent).", "weight": 8},
    {"id": 94, "category": "Compliance & Regulatory", "question": "How do you comply with the CAA (Consolidated Appropriations Act) plan transparency and gag clause prohibitions?", "evaluation": "CAA compliance is federally required. Ensure contract explicitly permits data sharing with consultants.", "weight": 9},
    {"id": 95, "category": "Compliance & Regulatory", "question": "How does your contract comply with findings from the FTC's 2024 Interim Report on PBM practices?", "evaluation": "The FTC identified rebate retention, spread pricing, and formulary manipulation as key concerns. Seek explicit compliance commitment.", "weight": 9},
    {"id": 96, "category": "Compliance & Regulatory", "question": "Describe your ERISA compliance processes, including claims and appeals procedures.", "evaluation": "Full ERISA compliance is required. Ask for summary of ERISA-compliant claims procedures.", "weight": 7},

    # Vendor Stability & References
    {"id": 97,  "category": "Vendor Stability & References", "question": "Describe your organizational structure, ownership, and any recent or planned mergers or acquisitions.", "evaluation": "Assess business continuity risk. Major ownership changes can disrupt service and contract terms.", "weight": 7},
    {"id": 98,  "category": "Vendor Stability & References", "question": "Provide three client references from employer groups of comparable size and industry.", "evaluation": "References should be verifiable and from comparable plans. Ask specifically about pricing, service, and audit results.", "weight": 8},
    {"id": 99,  "category": "Vendor Stability & References", "question": "What is your client retention rate over the past 3 years?", "evaluation": "Strong retention (90%+) indicates client satisfaction. Low retention is a red flag.", "weight": 7},
    {"id": 100, "category": "Vendor Stability & References", "question": "Are any services subcontracted to third parties? Identify subcontractors and their roles.", "evaluation": "Undisclosed subcontracting can affect service quality and accountability. All material relationships should be disclosed.", "weight": 6},

    # Transition & Implementation
    {"id": 101, "category": "Transition & Implementation", "question": "Describe your standard implementation plan, timeline, and key milestones.", "evaluation": "Standard: 90–120 day implementation. Ask for detailed project plan with named project manager.", "weight": 7},
    {"id": 102, "category": "Transition & Implementation", "question": "What member communication support do you provide during transition? Include sample materials.", "evaluation": "Member ID cards, formulary explanation, and pharmacy finder tools at least 30 days before go-live.", "weight": 6},
    {"id": 103, "category": "Transition & Implementation", "question": "How do you handle continuity of care for members on specialty medications during transition?", "evaluation": "Continuity of care provisions should prevent disruption for members on critical specialty therapies.", "weight": 8},
    {"id": 104, "category": "Transition & Implementation", "question": "Describe prior authorization transition procedures — will existing PAs be honored?", "evaluation": "Best practice: honor all active PAs for 90 days post-transition. New PA requirements during transition are disruptive.", "weight": 7},
    {"id": 105, "category": "Transition & Implementation", "question": "What is your process for historical claims data transfer from the incumbent PBM?", "evaluation": "Historical claim data portability is essential for continuity. Seek contractual commitment to claims data transfer.", "weight": 7},

    # Innovation & Value-Add
    {"id": 106, "category": "Innovation & Value-Add", "question": "Describe any manufacturer direct contracting or value-based arrangements for high-cost drugs.", "evaluation": "Direct contracts and outcomes-based arrangements can significantly reduce specialty costs.", "weight": 7},
    {"id": 107, "category": "Innovation & Value-Add", "question": "What cost containment programs do you offer for GLP-1 and weight loss medications?", "evaluation": "GLP-1 costs are rising rapidly. Look for coverage management, prior authorization, and outcomes programs.", "weight": 8},
    {"id": 108, "category": "Innovation & Value-Add", "question": "Describe your biosimilar pipeline strategy and estimated savings projections for the next 3 years.", "evaluation": "Major biosimilars entering market represent significant savings opportunity. Seek concrete projections.", "weight": 8},
    {"id": 109, "category": "Innovation & Value-Add", "question": "What analytics and predictive modeling capabilities do you offer for trend management?", "evaluation": "Predictive modeling can identify high-risk members and intervention opportunities. Ask for sample analyses.", "weight": 6},
    {"id": 110, "category": "Innovation & Value-Add", "question": "Do you offer international pharmacy or alternative sourcing programs for eligible drugs?", "evaluation": "International sourcing can reduce costs for certain brand drugs. Assess legal and safety compliance.", "weight": 6},

    # Specialty – Advanced
    {"id": 111, "category": "Specialty – Advanced", "question": "Provide specialty drug rebate pass-through percentages by therapeutic category (oncology, immunology, neurology, etc.).", "evaluation": "Category-level transparency allows better negotiation. Oncology rebates are often the largest pool.", "weight": 8},
    {"id": 112, "category": "Specialty – Advanced", "question": "Describe your accreditation status for specialty pharmacy (URAC, ACHC, or NCQA).", "evaluation": "URAC or ACHC accreditation indicates quality management compliance for specialty dispensing.", "weight": 7},
    {"id": 113, "category": "Specialty – Advanced", "question": "What is your clinical intervention rate for specialty patients and what cost savings do interventions generate?", "evaluation": "High-value specialty programs should demonstrate measurable ROI, typically $3–$5 per $1 invested.", "weight": 7},
    {"id": 114, "category": "Specialty – Advanced", "question": "How do you manage patient assistance programs (PAP) and 340B drug access for eligible plan members?", "evaluation": "PAP and 340B access can significantly offset costs for certain patient populations.", "weight": 7},
    {"id": 115, "category": "Specialty – Advanced", "question": "Describe your hub services model for specialty drug initiation, PA management, and adherence monitoring.", "evaluation": "Integrated hub services improve access and adherence for specialty patients, reducing total cost of care.", "weight": 6},
    {"id": 116, "category": "Specialty – Advanced", "question": "How do you manage white bagging vs. brown bagging policies for oncology and infusion drugs?", "evaluation": "White-bagging (PBM-dispensed) vs. brown-bagging affects cost and safety for infusion drugs.", "weight": 7},

    # Additional high-value questions
    {"id": 117, "category": "Rebates & Pass-Through", "question": "Describe how you handle manufacturer Patient Support Programs and ensure rebate eligibility is not jeopardized.", "evaluation": "Manufacturer support programs and rebate eligibility can conflict. Seek clear policies that maximize both.", "weight": 7},
    {"id": 118, "category": "Audit Rights", "question": "Can the plan sponsor conduct an audit covering all contract years retroactively, or only prospectively?", "evaluation": "Retroactive audit rights are essential for recovering historical underpayments.", "weight": 8},
    {"id": 119, "category": "Performance Guarantees", "question": "Are pharmacy network performance guarantees (e.g., generic fill rate, network adequacy) included?", "evaluation": "Network performance guarantees ensure the PBM maintains adequate pharmacy access and generic dispensing.", "weight": 7},
    {"id": 120, "category": "Administrative & Financial", "question": "What is your process for handling drug recalls, including obligations to notify the plan and members?", "evaluation": "Drug recall notification should be automatic and include member outreach for affected prescriptions.", "weight": 6},
    {"id": 121, "category": "Compliance & Regulatory", "question": "Describe how your contract complies with requirements under the Inflation Reduction Act affecting drug pricing.", "evaluation": "IRA provisions (Medicare price negotiation, inflation rebates) affect PBM economics. Seek pass-through provisions.", "weight": 8},
    {"id": 122, "category": "Pricing – Brand Drugs", "question": "How are brand drugs newly launching above WAC priced in your network, and are discounts guaranteed at launch?", "evaluation": "New brand drug launches often lack established AWP discounts. Seek guaranteed minimum discounts from launch date.", "weight": 7},
]


# ── Core export function ──────────────────────────────────────────────────────

def generate_rfp_export(analysis: PBMAnalysisReport) -> bytes:
    """Generate a prioritized RFP question bank XLSX based on the contract analysis. Returns raw bytes."""
    client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))
    mc = analysis.market_comparison
    pt = analysis.pricing_terms

    analysis_context = (
        f"Contract Grade: {analysis.overall_grade}\n\n"
        f"KEY CONCERNS:\n" + "\n".join(f"- {c}" for c in analysis.key_concerns) + "\n\n"
        f"PRICING ASSESSMENTS:\n"
        f"- Brand Retail: {mc.brand_retail_assessment}\n"
        f"- Generic Retail: {mc.generic_retail_assessment}\n"
        f"- Specialty: {mc.specialty_assessment}\n\n"
        f"HIGH-RISK AREAS:\n" +
        "\n".join(f"- {r.area}: {r.description}" for r in analysis.cost_risk_areas if r.risk_level == "high") + "\n\n"
        f"MEDIUM-RISK AREAS:\n" +
        "\n".join(f"- {r.area}: {r.description}" for r in analysis.cost_risk_areas if r.risk_level == "medium") + "\n\n"
        f"PRICING TERMS:\n"
        f"- Brand Retail AWP: {pt.brand_retail_awp_discount}\n"
        f"- Generic Retail AWP: {pt.generic_retail_awp_discount}\n"
        f"- Specialty AWP: {pt.specialty_awp_discount}\n"
        f"- Rebate: {pt.rebate_guarantee}\n"
        f"- MAC: {pt.mac_pricing_terms}\n"
        f"- Admin Fees: {pt.admin_fees}"
    )

    question_list = "\n".join(
        f"ID {q['id']} [{q['category']}]: {q['question']}"
        for q in QUESTION_BANK
    )

    prompt = (
        f"You are a PBM RFP specialist. Given a contract analysis with specific weaknesses, "
        f"select the 55 most important RFP questions for this client.\n\n"
        f"CONTRACT ANALYSIS:\n{analysis_context}\n\n"
        f"AVAILABLE QUESTIONS:\n{question_list}\n\n"
        f"Return a JSON array of exactly 55 objects. Sort them: HIGH priority first, then MEDIUM, then STANDARD.\n"
        f"Format:\n"
        f'[{{"id": 1, "priority": "HIGH", "contract_note": "why this matters for this contract"}}, ...]\n\n'
        f"Priority definitions:\n"
        f"- HIGH: Directly addresses an identified concern or below-market term in this contract\n"
        f"- MEDIUM: General risk area or moderate concern\n"
        f"- STANDARD: Good practice for all RFPs\n\n"
        f"Return ONLY the JSON array — no preamble or text outside the array."
    )

    response = client.messages.create(
        model="claude-opus-4-6",
        max_tokens=4000,
        messages=[{"role": "user", "content": prompt}],
    )

    priorities: dict[int, dict] = {}
    try:
        raw = response.content[0].text.strip()
        match = re.search(r"```(?:json)?\s*([\s\S]+?)\s*```", raw)
        if match:
            raw = match.group(1)
        for item in json.loads(raw):
            priorities[item["id"]] = item
    except Exception as e:
        print(f"[RFP] Failed to parse Claude prioritization: {e}")

    return _build_xlsx(analysis, priorities)


# ── XLSX builder ──────────────────────────────────────────────────────────────

def _build_xlsx(analysis: PBMAnalysisReport, priorities: dict[int, dict]) -> bytes:
    NAVY = "1E3A5F"
    HIGH_FILL = "FEE2E2"
    MED_FILL  = "FEF3C7"
    STD_FILL  = "EFF6FF"

    wb = openpyxl.Workbook()

    # ── Instructions sheet ───────────────────────────────────────────────────
    cover = wb.active
    cover.title = "Instructions"
    cover.column_dimensions["A"].width = 90

    cover_content = [
        (f"PBM RFP Question Bank — {date.today().strftime('%B %d, %Y')}", Font(bold=True, size=14, color=NAVY)),
        (f"Contract Grade: {analysis.overall_grade}  |  Parties: {analysis.contract_overview.parties}", Font(size=11, color="64748B")),
        ("", None),
        ("HOW TO USE THIS DOCUMENT", Font(bold=True, size=12, color=NAVY)),
        ("", None),
        ("1. Go to the 'RFP Questions' tab to find all selected questions.", Font(size=11)),
        ("2. Questions are sorted HIGH priority first, then MEDIUM, then STANDARD.", Font(size=11)),
        ("3. HIGH priority questions directly address weaknesses found in your current contract.", Font(size=11)),
        ("4. The 'Contract-Specific Note' column explains why each HIGH/MEDIUM question matters here.", Font(size=11)),
        ("5. Use the Evaluation Criteria column to score each PBM's response (1–10).", Font(size=11)),
        ("", None),
        ("KEY CONTRACT CONCERNS (address these in your RFP):", Font(bold=True, size=11, color="DC2626")),
    ]
    for concern in analysis.key_concerns:
        cover_content.append((f"  •  {concern}", Font(size=11)))
    cover_content += [
        ("", None),
        ("This document was generated by PBM Contract Analyzer. Review with your benefits counsel before use.",
         Font(size=10, color="94A3B8", italic=True)),
    ]

    for row_idx, (text, font) in enumerate(cover_content, start=1):
        cell = cover.cell(row=row_idx, column=1, value=text)
        if font:
            cell.font = font
        cell.alignment = Alignment(wrap_text=True)
        cover.row_dimensions[row_idx].height = 20

    # ── Questions sheet ──────────────────────────────────────────────────────
    ws = wb.create_sheet("RFP Questions")

    headers = ["#", "Priority", "Category", "Question", "Evaluation Criteria", "Weight", "Contract-Specific Note"]
    ws.append(headers)

    header_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = header_align

    ws.row_dimensions[1].height = 28
    col_widths = [5, 12, 24, 58, 48, 8, 52]
    for col, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Build prioritized list
    priority_order = {"HIGH": 0, "MEDIUM": 1, "STANDARD": 2}
    selected: list[dict] = []
    for q in QUESTION_BANK:
        if q["id"] in priorities:
            p = priorities[q["id"]]
            selected.append({
                **q,
                "priority": p.get("priority", "STANDARD"),
                "contract_note": p.get("contract_note", ""),
            })
    selected.sort(key=lambda x: priority_order.get(x.get("priority", "STANDARD"), 3))

    for row_num, q in enumerate(selected, start=2):
        priority = q.get("priority", "STANDARD")
        fill_color = HIGH_FILL if priority == "HIGH" else (MED_FILL if priority == "MEDIUM" else STD_FILL)
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        row_vals = [
            q["id"], priority, q["category"], q["question"],
            q["evaluation"], q["weight"], q.get("contract_note", ""),
        ]
        ws.append(row_vals)

        for col in range(1, 8):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            bold = col == 2  # bold the Priority column
            cell.font = Font(name="Calibri", bold=bold, size=10)

        ws.row_dimensions[row_num].height = 50

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
